# from openpyxl import Workbook, load_workbook
#
# # Load the existing workbook
# wb = load_workbook(filename='test.xlsx')
# # print(wb.sheetnames)
# # Select the target worksheet
# ws = wb["Sheet"]
#
# # что
# source_range = ws['A1:D1']
#
# # куда
# target_row = ws['2']
#
# # Copy the cell values and formatting using the copy() method
# for row in source_range:
#     for cell in row:
#         new_cell = target_row[cell.column_letter]
#         new_cell._style = cell._style
#         new_cell.value = cell.value if not cell.is_merged else None
#         if cell.is_merged:
#             range_string = cell.coordinate + ":" + cell.ending_coordinate
#             ws.unmerge_cells(range_string)
#             ws.merge_cells(range_string.replace(str(cell.row), str(target_row.row)))
#             new_cell._merged = True
#
# # Save the workbook
# wb.save(filename='your_file.xlsx')
#

from copy import copy
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from copy import copy
# def copy_cell(src_sheet, src_row, src_col,
#               tgt_sheet, tgt_row, tgt_col,
#               copy_style=True):
#     cell = src_sheet.cell(src_row, src_col)
#     new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
#     if cell.has_style and copy_style:
#         new_cell._style = copy(cell._style)
#
# filename = r"test.xlsx"
# wb = load_workbook(filename)
# print(wb.sheetnames)
# ws1 = wb['Sheet']
#
# filename = r"шапка.xlsx"
# wb2 = load_workbook(filename)
# print(wb2.sheetnames)
# ws2 = wb2['Лист1']
#
#
# ws1_last_row = ws1.max_row
#
# for i,row in enumerate(ws2.iter_rows(min_row=2, max_row=2), 1):
#     for cell in row:
#         ...
#         copy_cell(ws2, cell.row, cell.column,
#                   ws1, ws1_last_row+i, cell.column)

# wb.save(r"test2.xlsx")



wb1 = load_workbook('test.xlsx')
sheet1 = wb1['Sheet']
sheet1.insert_rows(0, 10, translate=True)
# sheet1.move_range("A1:R9", rows=10)

wb = load_workbook('шапка.xlsx')
shapka = wb['Sheet']

for i in range(1, 10):
    for j in range(1, 5):
        if isinstance(shapka.cell(row=i,column=j), MergedCell) or isinstance(sheet1.cell(row=i,column=j), MergedCell) :
            continue
        sheet1.cell(row=i,column=j).value = shapka.cell(row=i,column=j).value
        sheet1.cell(row=i, column=j).font =copy(shapka.cell(row=i, column=j).font)
        sheet1.cell(row=i, column=j).border =copy(shapka.cell(row=i, column=j).border)
        sheet1.cell(row=i, column=j).fill =copy(shapka.cell(row=i, column=j).fill)
        sheet1.cell(row=i, column=j).number_format =copy(shapka.cell(row=i, column=j).number_format)
        sheet1.cell(row=i, column=j).protection =copy(shapka.cell(row=i, column=j).protection)
        sheet1.cell(row=i, column=j).alignment =copy(shapka.cell(row=i, column=j).alignment)


wb1.save('test2.xlsx')

# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# # создадим произвольные данные
# data = [[row*col for col in range(1, 16)] for row in range(1, 31)]
# # добавляем данные на активный лист
# for row in data:
#     ws.append(row)
# # вставим 3 новые строки перед
# # существующей 7-ой строкой
# ws.insert_rows(7, 3)
# # сохраняем и смотрим
# wb.save('test3.xlsx')