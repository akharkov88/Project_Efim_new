import traceback
from copy import copy

from typing import (
    List,
    Optional,
)

from fastapi import (
    Depends,
    HTTPException,
    status,
)
from sqlalchemy.orm import Session

import models
import tables
import os
import pandas as pd
import pathlib
from services.auth import  get_session
from fastapi.encoders import jsonable_encoder
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
# from database import get_current_user
#
# print(get_current_user)
class BaseFileServices:
    def __init__(self, session: Session = Depends(get_session)):
        self.session = session

    def unload_file(self, file,
                    user: tables.User,
                    ) -> str:
        filename = self.create_uuid_table(file.filename, user)
        with open(pathlib.PurePath(pathlib.Path(os.path.abspath(os.getcwd())), "src", "file", filename), "wb") as f:
            f.write(file.file.read())
        return filename

    def create_uuid_table(self, name_fail,user: tables.User,) -> tables.BaseFile:
            try:
                operation = tables.BaseFile(
                    name=name_fail,
                    user_name=user.username,

                )
                self.session.add(operation)
                self.session.commit()
                return str(operation.id)
            except:
                print(traceback.format_exc())
                raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Ошибка создания файла")
                # raise JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={'message': "Уже существует запись"})

            # return "operation"

    def get_name_file(self,id):
        try:
            operation = (
                self.session
                .query(tables.BaseFile)
                .filter(
                    tables.BaseFile.id == id
                )
                .first()
            )
            #посмотреть какой запрос получится
            print(operation.compile(compile_kwargs={"literal_binds":True}))

            if not operation:
                raise HTTPException(status.HTTP_404)
            # return jsonable_encoder(operation)
            return operation.name
        except:
            print(traceback.format_exc())
            raise HTTPException(status.HTTP_400_BAD_REQUEST)
            # raise JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={'message': "Уже существует запись"})


    def reports_PTO(self,table:str):
        import uuid
        uuid = str(uuid.uuid4())
        name_fail=uuid.replace("-","_")

        table = table.tables.replace('colspan="100%"', '')
        table = table.replace('\n                             ', '')
        table = table.replace('\n                    ', '')

        workbook = Workbook()
        sheet = workbook.active
        with_column = {}
        soup = BeautifulSoup(table, 'lxml')
        if soup.find_all('table').__len__()==0:
            table=soup
        else:
            table = soup.find_all('table')[0]
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        try:
            row_marker = 1
            for row in table.find_all('tr'):
                column_marker = 1
                background_color = ""
                if row.get("style") and row.get("style").find("background-color") != -1:
                    background_color = row.get("style")[row.get("style").find("background-color: ") + 19:]
                columns = row.find_all('td')
                for column in columns:
                    if column.get("class")==None or "tehstolbec" not in column.get("class"):
                        if column.datalist:
                            column.datalist.decompose()
                        text_value=""
                        if len(str(column.get_text()).replace(" ","").replace("\n",""))==0:
                            if column.div:
                                if column.div.input:
                                    text_value=column.div.input.get("value")
                            if column.input:
                                if column.input.get("value")!=None:
                                    if column.input.get("value")!="":
                                        text_value = column.input.get("value")
                                        # background_color
                                        if str(column.input.get("id")).find("zagolovok_")!=-1:
                                            background_color="a94442"
                        else:
                            if column.div:
                                column.div.decompose()
                            text_value=column.get_text()

                        if len(text_value)==4:
                            pass


                        while isinstance(sheet.cell(row=row_marker, column=column_marker), MergedCell):
                            column_marker += 1
                        sheet.cell(row=row_marker, column=column_marker, value=text_value)
                        sheet.cell(row=row_marker, column=column_marker).border = thin_border
                        sheet.cell(row=row_marker, column=column_marker).alignment = Alignment(horizontal='center')
                        text_len = len(text_value)
                        if text_len != 0 and with_column.setdefault(get_column_letter(column_marker), text_len + 2):
                            if with_column[get_column_letter(column_marker)] <= text_len + 2:
                                with_column[get_column_letter(column_marker)] = text_len + 2
                                sheet.column_dimensions[get_column_letter(column_marker)].width = text_len + 2
                        if background_color != "":
                            try:
                                sheet[row_marker][column_marker - 1].fill = PatternFill(start_color=background_color,
                                                                                        end_color=background_color,
                                                                                        fill_type="solid")
                            except:
                                print(traceback.format_exc())
                                print("error",row_marker,column_marker - 1)

                        if column.get("colspan") or column.get("rowspan"):

                            merge_column = column_marker
                            merge_row = row_marker
                            if column.get("colspan"):
                                print(column.get("colspan"))
                                merge_column += int(column.get("colspan")) - 1
                            if column.get("rowspan"):
                                merge_row += int(column.get("rowspan")) - 1

                            sheet.merge_cells(start_row=row_marker, start_column=column_marker, end_row=int(merge_row),
                                              end_column=int(merge_column))
                        if column.get("colspan"):
                            column_marker += int(column.get("colspan")) - 1

                        column_marker += 1
                row_marker += 1
            workbook.save(filename=pathlib.PurePath(pathlib.Path(os.path.abspath(os.getcwd())), "src", "file", name_fail))
        except:
            print(traceback.format_exc())
            workbook.save(filename="converter_html_in_xlsx.xlsx")

        return name_fail


    def reports_PTO_KP(self,table:str):
        import uuid
        uuid = str(uuid.uuid4())
        name_fail=uuid.replace("-","_")

        table = table.tables.replace('colspan="100%"', '')
        table = table.replace('\n                             ', '')
        table = table.replace('\n                    ', '')

        workbook = Workbook()
        sheet = workbook.active
        from openpyxl import load_workbook
        wb = load_workbook('шапка.xlsx')
        shapka = wb['Sheet']

        for i in range(1, shapka.max_row+2
                       ):
            for j in range(1, shapka.max_column+2):
                if isinstance(shapka.cell(row=i, column=j), MergedCell) or isinstance(sheet.cell(row=i, column=j),
                                                                                      MergedCell):
                    continue
                sheet.cell(row=i, column=j).value = shapka.cell(row=i, column=j).value
                sheet.cell(row=i, column=j).font = copy(shapka.cell(row=i, column=j).font)
                sheet.cell(row=i, column=j).border = copy(shapka.cell(row=i, column=j).border)
                sheet.cell(row=i, column=j).fill = copy(shapka.cell(row=i, column=j).fill)
                sheet.cell(row=i, column=j).number_format = copy(shapka.cell(row=i, column=j).number_format)
                sheet.cell(row=i, column=j).protection = copy(shapka.cell(row=i, column=j).protection)
                sheet.cell(row=i, column=j).alignment = copy(shapka.cell(row=i, column=j).alignment)

        with_column = {}

        soup = BeautifulSoup(table, 'lxml')
        if soup.find_all('table').__len__()==0:
            table=soup
        else:
            table = soup.find_all('table')[0]
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        try:
            row_marker = shapka.max_row+4
            for row in table.find_all('tr'):
                column_marker = 1
                background_color = ""
                if row.get("style") and row.get("style").find("background-color") != -1:
                    background_color = row.get("style")[row.get("style").find("background-color: ") + 19:]
                columns = row.find_all('td')
                for column in columns:
                    if column.get("class")==None or "tehstolbec" not in column.get("class"):
                        if column.datalist:
                            column.datalist.decompose()
                        text_value=""
                        if len(str(column.get_text()).replace(" ","").replace("\n",""))==0:
                            if column.div:
                                if column.div.input:
                                    text_value=column.div.input.get("value")
                            if column.input:
                                if column.input.get("value")!=None:
                                    if column.input.get("value")!="":
                                        text_value = column.input.get("value")
                                        # background_color
                                        if str(column.input.get("id")).find("zagolovok_")!=-1:
                                            background_color="a94442"
                        else:
                            if column.div:
                                column.div.decompose()
                            text_value=column.get_text()

                        if len(text_value)==4:
                            pass


                        while isinstance(sheet.cell(row=row_marker, column=column_marker), MergedCell):
                            column_marker += 1
                        sheet.cell(row=row_marker, column=column_marker, value=text_value)
                        sheet.cell(row=row_marker, column=column_marker).border = thin_border
                        sheet.cell(row=row_marker, column=column_marker).alignment = Alignment(horizontal='center')
                        text_len = len(text_value)
                        if text_len != 0 and with_column.setdefault(get_column_letter(column_marker), text_len + 2):
                            if with_column[get_column_letter(column_marker)] <= text_len + 2:
                                with_column[get_column_letter(column_marker)] = text_len + 2
                                sheet.column_dimensions[get_column_letter(column_marker)].width = text_len + 2
                        if background_color != "":
                            try:
                                sheet[row_marker][column_marker - 1].fill = PatternFill(start_color=background_color,
                                                                                        end_color=background_color,
                                                                                        fill_type="solid")
                            except:
                                print(traceback.format_exc())
                                print("error",row_marker,column_marker - 1)


                        if column.get("colspan") or column.get("rowspan"):

                            merge_column = column_marker
                            merge_row = row_marker
                            if column.get("colspan"):
                                print(column.get("colspan"))
                                merge_column += int(column.get("colspan")) - 1
                            if column.get("rowspan"):
                                merge_row += int(column.get("rowspan")) - 1

                            sheet.merge_cells(start_row=row_marker, start_column=column_marker, end_row=int(merge_row),
                                              end_column=int(merge_column))
                        if column.get("colspan"):
                            column_marker += int(column.get("colspan")) - 1

                        column_marker += 1
                row_marker += 1
            workbook.save(filename=pathlib.PurePath(pathlib.Path(os.path.abspath(os.getcwd())), "src", "file", name_fail))
        except:
            print(traceback.format_exc())
            workbook.save(filename="converter_html_in_xlsx.xlsx")


        # workbook.save(filename="test.xlsx")
        #
        # load_workbook = load_workbook("test.xlsx")
        # load_workbook_ws1 = load_workbook['Sheet1']
        #
        # shablon_shapka = load_workbook(r"/home/aleksey/pythonProject/Project_Efim_new/Shablon_excel/шапка.xlsx")
        # shablon_shapka_ws1 = shablon_shapka['Sheet1']
        #
        # _ = [load_workbook_ws1.append(row) for row in shablon_shapka_ws1.iter_rows(min_row=1, max_row=9)]
        #
        # load_workbook.save(filename="test2.xlsx")
        return name_fail
