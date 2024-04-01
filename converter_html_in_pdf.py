import traceback

from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from fpdf import FPDF


ff = '<table class="first iksweb" id="first_table">                        <tbody id="earthWorks_1">                    <tr><td colspan="6" style="padding: 0;"><input id="zagolovok_1" style="width: 100%; height: 100%; border:  0;" autocomplete="off" value="23">                               </td><td colspan="4" class="tehstolbec"><button id="del_razdel_1" type="button" class="btn-primary"> Удалить раздел</button></td>                    </tr>                        <tr style="background-color: #fce1b3">                            <td rowspan="2" class="tehstolbec">Тех столбец</td>                            <td>№</td>                            <td rowspan="2" style="font-weight: bold;">Наименование</td>                            <td colspan="4" style="font-weight: bold;">Стоимость</td>                            <td colspan="4" style="font-weight: bold;">Справочная информация</td>                            <td colspan="8" style="font-weight: bold;" class="pto">Аналитический блок</td>                        </tr>                        <tr style="background-color: #fce1b3">                            <td style="font-weight: bold;">п/п</td>                            <td style="font-weight: bold;">Ед.изм.</td>                            <td style="font-weight: bold;">Кол-во</td>                            <td style="font-weight: bold;">Цена</td>                            <td style="font-weight: bold;">Всего</td>                            <td rowspan="1" style="font-weight: bold;">Тр-затр (чел-дн)</td>                            <td rowspan="1" style="font-weight: bold;">Столбец для анализа</td>                            <td rowspan="1" style="font-weight: bold;"></td>                            <td rowspan="1" style="font-weight: bold;">Примечание</td>                            <td style="font-weight: bold;" class="pto">Кол-во</td>                            <td style="font-weight: bold;" class="pto">Цена руб-ед</td>                            <td style="font-weight: bold;" class="pto">Всего</td>                            <td style="font-weight: bold;" class="pto">Тр-затр (чел-дн)</td>                            <td style="font-weight: bold;" class="pto">Стоимость дня рабочего</td>                        </tr>                        </tbody>                        <tbody id="earthWorks_costWork_1">                        <tr>                            <td class="tehstolbec"><button id="button_costWork_collapse_1" type="button" onclick="collaps_tbody("costWork", 1, this)">Свернуть</button></td>                            <td contenteditable="false" style="border-left: none; border-right: none;"></td>                            <td contenteditable="false" id="earthWorks_costWork_addLine_1" style="font-weight: bold; border-left: none;" colspan="4" onmouseover="showButton(event, this)" onmouseout="closeButton(event,this)">                            Стоимость работ                            <div style="display: none;"><button type="button" onclick="addLine(event,this)">addLine</button></div>                            </td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto"></td>                        </tr>                    </tbody><tbody id="earthWorks_costMaterial_1">                        <tr>                        <td class="tehstolbec"><button id="button_costMaterial_collapse_1" type="button" onclick="collaps_tbody("costMaterial", 1, this)">Свернуть</button></td>                            <td contenteditable="false" style="border-left: none; border-right: none;"></td>                            <td contenteditable="false" id="earthWorks_costMaterial_addLine_1" style="font-weight: bold; border-left: none;" colspan="4" onmouseover="showButton(event, this)" onmouseout="closeButton(event,this)">                            Стоимость материала                            <div style="display: none;"><button onclick="addLine(event,this)">addLine</button></div>                            </td>                            <td contenteditable="false"></td>                            <td contenteditable="false" colspan="2" style="font-weight: bold;"> Ед изм расчётные </td>                            <td contenteditable="false" style="font-weight: bold;"> Ед изм (М2) для окраски </td>                            <td contenteditable="false"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Закупка по счету - ед изм расчётные</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Остаток закупки</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">ед изм к закупке</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">руб-ед</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Счёта</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Срок поставки</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">ПРИМЕЧАНИЕ</td>                        </tr>                    </tbody>                    <tbody id="earthWorks_Special_1">                    <tr>                    <td class="tehstolbec"><button id="button_Special_collapse_1" type="button" onclick="collaps_tbody("Special", 1, this)">Свернуть</button></td>                            <td contenteditable="false" style="border-left: none; border-right: none;"></td>                             <td contenteditable="false" id="earthWorks_Special_addLine_1" style="font-weight: bold; border-left: none;" colspan="4" onmouseover="showButton(event, this)" onmouseout="closeButton(event,this)">                            Стоимость спецтехники                            <div style="display: none;"><button onclick="addLine(event,this)">addLine</button></div>                            </td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Стоимость аренды</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Телефон</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Счет</td>                            <td contenteditable="false" style="font-weight: bold;" class="pto">Срок</td>                        </tr>                    </tbody>                     <tbody id="earthWorks_Total_1">                        <tr>                        <td style="border-right: none;" class="tehstolbec"></td>                            <td contenteditable="false" style="border-left: none; border-right: none;"></td>                            <td contenteditable="false" style="font-weight: bold;border-left: none;" colspan="4"> Итого                            </td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false" style="" class="pto"></td>                            <td contenteditable="false" style="" class="pto"></td>                            <td contenteditable="false" style="" class="pto"></td>                            <td contenteditable="false" style="" class="pto"></td>                        </tr>                        <tr id="stroka">                        <td class="tehstolbec"></td>                            <td contenteditable="false">1</td>                            <td contenteditable="false" style="font-weight: bold;">                                Накладные расходы                            </td>                            <td contenteditable="false">%</td>                            <td contenteditable="true" class="percent"></td>                            <td contenteditable="true"></td>                            <td contenteditable="false"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                        </tr>                        <tr id="stroka">                        <td class="tehstolbec"></td>                            <td contenteditable="false">2</td>                            <td contenteditable="false" style="font-weight: bold;">                                Вспомогательные материалы                            </td>                            <td contenteditable="false">%</td>                            <td contenteditable="true" class="percent"></td>                            <td contenteditable="true"></td>                            <td contenteditable="false"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto" <="" td="">                            </td><td contenteditable="true" style="" class="costWork_role_boss pto"></td>                        </tr>                        <tr id="stroka">                        <td class="tehstolbec"></td>                            <td contenteditable="false">3</td>                            <td contenteditable="false" style="font-weight: bold;">                                Транспортные расходы                            </td>                            <td contenteditable="false">%</td>                            <td contenteditable="true" class="percent"></td>                            <td contenteditable="true"></td>                            <td contenteditable="false"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                            <td contenteditable="true" style="" class="costWork_role_boss pto"></td>                        </tr>                        <tr id="stroka">                        <td class="tehstolbec"></td>                            <td contenteditable="false">4</td>                            <td contenteditable="false" style="font-weight: bold; text-align: right">                                ВСЕГО:                            </td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                        </tr>                        <tr id="stroka">                        <td class="tehstolbec"></td>                            <td contenteditable="false">5</td>                            <td contenteditable="false" style="font-weight: bold;">                                В том числе НДС - 20% за нал от материалов                            </td>                            <td contenteditable="false">%</td>                            <td contenteditable="true" class="percent"></td>                            <td contenteditable="true"></td>                            <td contenteditable="false"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                            <td contenteditable="true"></td>                        </tr>                    </tbody>                    <tbody id="result_total">                    <tr class="blank_row">                    </tr>                        <tr>                        <td class="tehstolbec"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false" style="font-weight: bold; text-align: right">                                Итого по разделам:                            </td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                        </tr>                        <tr>                        <td class="tehstolbec"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false" style="font-weight: bold;">                               Итого НДС:                               </td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                            <td contenteditable="false"></td>                        </tr>                    </tbody>                    </table>'
ff=ff.replace('colspan="100%"','')
ff=ff.replace('\n                             ','')
ff=ff.replace('\n                    ','')


workbook = Workbook()
sheet = workbook.active
with_column={}


soup = BeautifulSoup(ff, 'lxml')
table = soup.find_all('table')[0]
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# class PDF(FPDF):
#
#     def set_style(self):
#         # цвет текста
#         self.set_text_color(255)
#         # цвет линий таблицы
#         self.set_draw_color(255, 0, 0)
#         # ширина линии
#         self.set_line_width(0.3)
#         # Восстановление цвета и шрифта:
#         self.set_fill_color(224, 235, 255)
#         self.set_text_color(0)
#     def add_row_table(self,text):
#         self.cell(42, 7, text, 1, 0, "C", True)
#         self.ln()
#
#     def colored_table(self, headings, rows, col_widths=(42, 39, 35, 42)):
#         self.set_fill_color(255, 100, 0)
#         # цвет текста
#         self.set_text_color(255)
#         # цвет линий таблицы
#         self.set_draw_color(255, 0, 0)
#         # ширина линии
#         self.set_line_width(0.3)
#         # жирный шрифт
#         # self.set_font(style="B")
#         for col_width, heading in zip(col_widths, headings):
#             self.cell(col_width, 7, heading, 1, 0, "C", True)
#         self.ln()
#         # Восстановление цвета и шрифта:
#         self.set_fill_color(224, 235, 255)
#         self.set_text_color(0)
#         # self.set_font()
#         fill = False
#         for row in rows:
#             self.cell(col_widths[0], 6, row[0], "LR", 0, "L", fill)
#             self.cell(col_widths[1], 6, row[1], "LR", 0, "L", fill)
#             self.cell(col_widths[2], 6, row[2], "LR", 0, "R", fill)
#             self.cell(col_widths[3], 6, row[3], "LR", 0, "R", fill)
#             self.ln()
#             fill = not fill
#         self.cell(sum(col_widths), 0, "", "T")
#

class Pareser_html(BeautifulSoup,FPDF):
    def __init__(self):
        FPDF.__init__(self,orientation = 'L', format = 'A4')
    def version1(self,text):
        rezultat=[]

        text = text.replace('colspan="100%"', '')
        text = text.replace('\n                             ', '')
        text = text.replace('\n                    ', '')

        workbook = Workbook()
        sheet = workbook.active
        with_column = {}

        soup = BeautifulSoup(text , 'lxml')
        table = soup.find_all('table')[0]
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        row_marker = 1

        for row in table.find_all('tr'):
            column_marker = 1
            print(row)
            background_color = ""
            if row.get("style") and row.get("style").find("background-color") != -1:
                background_color = row.get("style")[row.get("style").find("background-color: ") + 19:]

            rezultat.append({"row_marker":row_marker,"background_color":background_color,"value":[]})
            columns = row.find_all('td')
            for column in columns:
                if column.get("class") == None or "tehstolbec" not in column.get("class"):
                    print(column)

                    if column.datalist:
                        column.datalist.decompose()
                    text_value = ""
                    if len(str(column.get_text()).replace(" ", "").replace("\n", "")) == 0:
                        if column.div:
                            if column.div.input:
                                text_value = column.div.input.get("value")
                    else:
                        if column.div:
                            column.div.decompose()
                        text_value = column.get_text()


                    while isinstance(sheet.cell(row=row_marker, column=column_marker), MergedCell):
                        column_marker += 1
                    # sheet.cell(row=row_marker, column=column_marker, value=text_value)
                    # sheet.cell(row=row_marker, column=column_marker).border = thin_border
                    # sheet.cell(row=row_marker, column=column_marker).alignment = Alignment(horizontal='center')
                    rezultat[len(rezultat)-1]["value"].append({"text_value":text_value,
                                   "merge_column":1,"merge_row":1,"row_marker":row_marker,"column_marker":column_marker

                                     })

                    text_len = len(text_value)
                    if text_len != 0 and with_column.setdefault(get_column_letter(column_marker), text_len + 2):
                        if with_column[get_column_letter(column_marker)] <= text_len + 2:
                            with_column[get_column_letter(column_marker)] = text_len + 2
                            sheet.column_dimensions[get_column_letter(column_marker)].width = text_len + 2
                    # if background_color != "":
                    #     sheet[row_marker][column_marker - 1].fill = PatternFill(start_color=background_color,
                    #                                                             end_color=background_color,
                    #                                                             fill_type="solid")

                    if column.get("colspan") or column.get("rowspan"):

                        merge_column = column_marker
                        merge_row = row_marker
                        if column.get("colspan"):
                            print(column.get("colspan"))
                            merge_column += int(column.get("colspan")) - 1
                        if column.get("rowspan"):
                            merge_row += int(column.get("rowspan")) - 1


                        rezultat[len(rezultat) - 1]["value"][len(rezultat[len(rezultat) - 1]["value"])-1]["merge_column"]= int(merge_column)
                        rezultat[len(rezultat) - 1]["value"][len(rezultat[len(rezultat) - 1]["value"])-1]["merge_row"]= int(merge_row)

                        #

                    if column.get("colspan"):
                        column_marker += int(column.get("colspan")) - 1
                    # if column.get("rowspan"):
                    #     row_marker += int(column.get("rowspan"))-1

                    column_marker += 1
            row_marker += 1
        return rezultat
    # def excel(self, rezult):
    def parse_in_excel(self,rezult):
        workbook = Workbook()
        sheet = workbook.active

        # column_marker=1
        # row_marker=1
        for row in rezult:
            print(row)
            for column in row["value"]:
                print(column)

                while isinstance(sheet.cell(row=column["row_marker"], column=column["column_marker"]), MergedCell):
                    column["column_marker"] += 1
                sheet.cell(row=column["row_marker"], column=column["column_marker"], value=column["text_value"])
                sheet.cell(row=column["row_marker"], column=column["column_marker"]).border = thin_border
                sheet.cell(row=column["row_marker"], column=column["column_marker"]).alignment = Alignment(horizontal='center')
                if row["background_color"]!="":
                    sheet[column["row_marker"]][column["column_marker"] - 1].fill = PatternFill(start_color=row["background_color"],
                                                                            end_color=row["background_color"],
                                                                        fill_type="solid")
                if int(column["merge_column"])>1 and int(column["merge_row"])>1:
                    sheet.merge_cells(start_row=column["row_marker"], start_column=column["column_marker"], end_row=int(column["merge_row"]),
                                      end_column=int(column["merge_column"]))
                if sheet.column_dimensions[get_column_letter(column["column_marker"])].width<len(column["text_value"]):
                    sheet.column_dimensions[get_column_letter(column["column_marker"])].width = len(column["text_value"])
        workbook.save(filename="converter_html_in_xlsx.xlsx")

    def parse_in_pdf(self,rezult):
        # pdf =FPDF()
        # super(FPDF, self).__init__()
        self.add_font("Sans", style="", fname="src/static/main/assets/fonts/ofont.ru_Noto Sans.ttf", uni=True)
        self.add_font("Sans", style="B", fname="src/static/main/assets/fonts/ofont.ru_Noto Sans.ttf", uni=True)
        self.add_font("Sans", style="I", fname="src/static/main/assets/fonts/ofont.ru_Noto Sans.ttf", uni=True)
        self.add_font("Sans", style="BI", fname="src/static/main/assets/fonts/ofont.ru_Noto Sans.ttf", uni=True)
        # настройка шрифта
        self.set_font("Sans", size=8)
        self.add_page()
        self.set_style()

        # i=0
        # j=0
        # bor_val_column=[]
        # for row in rezult:
        #     bor_val_column.append([])
        #     for col in row["value"]:
        #         bor_val_column[i][j]={"text":col["text_value"],"border":""}
        #         j += 1
        #     i += 1

        with_column={}
        i=1
        for row in rezult:
            for col in row["value"]:
                col["text_value"]=col["text_value"].strip()
                if col["merge_column"]==1:
                    if with_column.get(str(i),False):
                        if len(col["text_value"])>with_column[str(i)]["len"]:
                            with_column[str(i)]["len"]=len(col["text_value"])
                            with_column[str(i)]["text"]=col["text_value"]

                    else:
                        with_column.setdefault(str(i),{"len":len(col["text_value"]),"text":col["text_value"]})
                    i+=1


            i=1

        analiz_merge_row={}
        i=1
        for row in rezult:
            for col in row["value"]:
                try:

                    if col["merge_row"] != 1:
                        analiz_merge_row[col["column_marker"]]=col["merge_row"]
                    #
                    # if analiz_merge_row.get(col["column_marker"],False):
                    #     if analiz_merge_row[col["column_marker"]]!=0:
                    #         analiz_merge_row[col["column_marker"]]-=1
                    #         self.add_row_table("", with_column.setdefault(str(i),{"len": len(col["text_value"]),"text": ""}),col["merge_row"])
                    #         i += 1
                    #         continue
                    print(col)

                    if col["merge_column"] == 1:
                        self.add_row_table(col["text_value"], with_column.setdefault(str(i),{"len": len(col["text_value"]),"text": col["text_value"]}),col["merge_row"])

                    else:
                        ff={"len":0,"text": ""}
                        for j in range(0,col["merge_column"]):
                            if with_column.get(str(i), True):
                                with_column.setdefault(str(i),{"len": len(col["text_value"]),"text": col["text_value"]})

                            ff["len"]+=with_column[str(i)]["len"]
                            ff["text"]+=with_column[str(i)]["text"]
                            i += 1

                        self.add_row_table(col["text_value"], ff,col["merge_row"])

                except:
                    print(traceback.format_exc())
                    pass
                i+=1

            # self.output("convertert_html_in_pdf.pdf")

            self.add_perenos_row_table()
            i=1

        # self.add_row_table("text",0)
        # self.add_perenos_row_table()
        # self.add_row_table("textdddddd",4)
        # self.add_row_table("text",0)
        # self.add_row_table("text",0)
        # self.add_perenos_row_table()
        # self.add_row_table("text",0)
        # self.add_row_table("text",0)
        # self.add_row_table("text",0)
        # self.add_row_table("text")
        # pdf.colored_table(["111111111111"], ["222222222","3333333333333"])
        self.output("convertert_html_in_pdf.pdf")
    def set_style(self):
        # цвет текста
        self.set_text_color(255)
        # цвет линий таблицы
        self.set_draw_color(255, 0, 0)
        # ширина линии
        self.set_line_width(0.3)
        # Восстановление цвета и шрифта:
        self.set_fill_color(224, 235, 255)
        self.set_text_color(0)
    def add_row_table(self,text,i,h):
        try:
            self.cell(i["len"]*1.8, 7, text, "TLRB", 0, "C", True)
        except:
            pass
    def add_perenos_row_table(self):
        self.ln()

    def colored_table(self, headings, rows, col_widths=(42, 39, 35, 42)):
        self.set_fill_color(255, 100, 0)
        # цвет текста
        self.set_text_color(255)
        # цвет линий таблицы
        self.set_draw_color(255, 0, 0)
        # ширина линии
        self.set_line_width(0.3)
        # жирный шрифт
        # self.set_font(style="B")
        for col_width, heading in zip(col_widths, headings):
            self.cell(col_width, 7, heading, 1, 0, "C", True)
        self.ln()
        # Восстановление цвета и шрифта:
        self.set_fill_color(224, 235, 255)
        self.set_text_color(0)
        # self.set_font()
        fill = False
        for row in rows:
            self.cell(col_widths[0], 6, row[0], "LR", 0, "L", fill)
            self.cell(col_widths[1], 6, row[1], "LR", 0, "L", fill)
            self.cell(col_widths[2], 6, row[2], "LR", 0, "R", fill)
            self.cell(col_widths[3], 6, row[3], "LR", 0, "R", fill)
            self.ln()
            fill = not fill
        self.cell(sum(col_widths), 0, "", "T")



# pdf = PDF()
Pareser_html = Pareser_html()
Pareser_html.parse_in_excel(Pareser_html.version1(ff))
Pareser_html.parse_in_pdf(Pareser_html.version1(ff))


